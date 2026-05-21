import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Influencers Barberos España"

headers = [
    "Tier", "Handle Instagram", "Nombre", "Seguidores IG", "TikTok/YT",
    "Ciudad", "Email", "Telefono / WhatsApp", "Web / Booksy",
    "Especialidad / Por que VISAI", "Nota Outreach", "Estado"
]

header_fill = PatternFill("solid", fgColor="222222")
tier1_fill = PatternFill("solid", fgColor="2d1b69")
tier2_fill = PatternFill("solid", fgColor="1b3a4b")
tier3_fill = PatternFill("solid", fgColor="1b2d1b")

header_font = Font(name="Calibri", bold=True, color="FFD700", size=11)
white_font = Font(name="Calibri", color="FFFFFF")
gold_bold = Font(name="Calibri", color="FFD700", bold=True)
green_font = Font(name="Calibri", color="90EE90")

data = [
    # MEGA
    ["MEGA", "@samuu.barberr", "Samuel Carballares", "~1.000.000", "3.1M TikTok | 1.5M YT",
     "Madrid (Arganda)", "samu.barber@muganagency.com", "-", "TikTok / Instagram DM",
     "El mas grande de España. Viral por cortes a homeless. Academia online.",
     "Contactar por agencia: muganagency.com. Historia VISAI = contenido perfecto.", ""],

    ["MEGA", "@ggsoaress", "Gabriel Soares", "648.000", "446K TikTok",
     "Barcelona (Olesa de Montserrat)", "gsoares.egworks@gmail.com", "-",
     "gabrielsoaresmenssalon.booksy.com",
     "Top barbero BCN. Mensstyle, tutoriales. Owner gsoaressalon + soaresmencare.",
     "Email directo disponible. Proponer colaboracion B2B.", ""],

    ["MEGA", "@artebydianatroyano", "Diana Troyano", "260.000", "197K TikTok",
     "Madrid (Chamberi)", "info@artebydianatroyano.es", "685 21 85 73 / 910 41 27 64",
     "artebydianatroyano.es",
     "COLORIMETRIA + VISAGISMO — especialidad identica a VISAI. Audiencia = nuestro target.",
     "PRIORITARIA. Email + telefono disponibles. Proponer colaboracion estrategica.", ""],

    ["MEGA", "@erickgomezz1", "Erick Gomez", "252.000", "YT propio",
     "Madrid (Leganes / Aluche)", "contacto@erickgomezacademy.com",
     "912 08 18 06 (Aluche) / 912 59 64 94 (Alto Extremadura)", "ballinbarbershop.com",
     "Ballin Barbershop + Ballin Academy. Gran comunidad de barberos.",
     "Email + telefono directos. Contactar para partnership academy.", ""],

    ["MEGA", "@sickerblack", "IKER Irigoyen", "218.000", "-",
     "Pamplona (Rotxapea)", "-", "-", "akerra.shop",
     "Barbero & Asesor de Imagen. CEO akerra.shop. Audiencia masculina premium.",
     "Solo DM Instagram o visita. C/Errotazar 21, Rotxapea, Pamplona.", ""],

    ["MEGA", "@ceache.barber", "Cesar Llamuca (Ceache)", "184.000", "-",
     "Barcelona (Nou Barris)", "-", "933 49 87 88 / 631 68 25 72", "chbarberstudio.com",
     "Barbero del Barca. Futbolistas como clientes. Varios estudios Ceache BCN.",
     "Telefono directo disponible. Imagen aspiracional + futbol = viral.", ""],

    ["MEGA", "@riduan_liindo", "Ridouan Bouzagou", "183.000", "-",
     "Barcelona (Mataro / Martorell)", "-", "-", "@riduan.liindo.studio (IG estudio)",
     "Barbero de Lamine Yamal y Nico Williams. VIP studio. Salio en Espejo Publico.",
     "El barbero mas viral de España ahora. DM Instagram unica via.", ""],

    # MID
    ["MID", "@sanrothebarber", "Raul Sanroma", "~40.000 est.", "Sanro The Barber YT",
     "Barcelona (El Prat de Llobregat)", "-", "-", "Reservas via IG",
     "Barbero del Barca (Gavi, Fermin, Lamine). Ambassador Booksy + Slick Gorilla.",
     "DM Instagram. C/Enric Morera 90, El Prat de Llobregat.", ""],

    ["MID", "@soherbarber", "Soher", "35.000", "-",
     "Barcelona", "-", "-", "@barber.cm (negocio)",
     "Top barbero BCN. Owner barber.cm. En ranking nacional.",
     "DM Instagram. Perfil solido para codigo local BCN.", ""],

    ["MID", "@maonzbarber", "Diego Velasco (Maonz)", "32.000", "YT + TikTok",
     "España", "-", "-", "Tienda online propia",
     "Educador online barberia. Cursos, tutoriales, tienda.",
     "DM Instagram. Audiencia de barberos = multiplicador de codigos.", ""],

    ["MID", "@ballin.barbershop", "Ballin Barbershop", "18.000", "-",
     "Madrid (Leganes)", "contacto@erickgomezacademy.com", "912 08 18 06", "ballinbarbershop.com",
     "Barberia de Erick Gomez. Cuenta institucional.",
     "Contactar via Erick directamente (ver fila MEGA).", ""],

    # LOCAL
    ["LOCAL", "@barberia_barcelona", "Barberia Barcelona", "~8.000 est.", "-",
     "Barcelona", "-", "-", "-",
     "Cuenta colectiva barberos BCN.",
     "Comunidad local. Util para awareness en BCN.", ""],

    ["LOCAL", "@madrid.shaving.club", "Madrid Shaving Club", "~5.000", "-",
     "Madrid (3 sedes)", "-", "-", "-",
     "Barberia premium Madrid. Sedes: Moncloa, Estrecho, Alonso Martinez.",
     "Cliente ideal VISAI: barberia premium multisede. Proponer codigo para sus barberos.", ""],
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, row_data in enumerate(data, 2):
    tier = row_data[0]
    fill = tier1_fill if tier == "MEGA" else (tier2_fill if tier == "MID" else tier3_fill)
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill
        if col_idx == 2:
            cell.font = gold_bold
        elif col_idx in (7, 8) and val and val != "-":
            cell.font = green_font  # highlight contacts we have
        else:
            cell.font = white_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

widths = [10, 22, 22, 14, 18, 22, 32, 28, 28, 45, 50, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 32
for r in range(2, len(data) + 2):
    ws.row_dimensions[r].height = 65

ws.freeze_panes = "A2"

# Resumen de contactos
ws3 = wb.create_sheet("Resumen Contactos")
ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 30
ws3.column_dimensions["D"].width = 20

ws3["A1"] = "Resumen contactos disponibles"
ws3["A1"].font = Font(bold=True, size=14, color="FFD700")
ws3["A1"].fill = PatternFill("solid", fgColor="1a1a2e")

ws3_headers = ["Nombre", "Email", "Telefono / WA", "Prioridad"]
for col, h in enumerate(ws3_headers, 1):
    cell = ws3.cell(row=2, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font

contacts = [
    ["Diana Troyano", "info@artebydianatroyano.es", "685 21 85 73", "1 - MAXIMA"],
    ["Erick Gomez", "contacto@erickgomezacademy.com", "912 08 18 06", "2 - ALTA"],
    ["Gabriel Soares", "gsoares.egworks@gmail.com", "-", "3 - ALTA"],
    ["Samuel Carballares", "samu.barber@muganagency.com", "-", "4 - ALTA (agencia)"],
    ["Ceache (Cesar)", "-", "933 49 87 88", "5 - ALTA"],
    ["IKER Irigoyen", "-", "-", "6 - Solo DM IG"],
    ["Ridouan Bouzagou", "-", "-", "7 - Solo DM IG"],
    ["Raul Sanroma", "-", "-", "8 - Solo DM IG"],
    ["Soher", "-", "-", "9 - Solo DM IG"],
    ["Maonz (Diego Velasco)", "-", "-", "10 - Solo DM IG"],
]

for row_idx, row in enumerate(contacts, 3):
    for col_idx, val in enumerate(row, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = PatternFill("solid", fgColor="111122")
        cell.font = green_font if (col_idx in (2, 3) and val != "-") else white_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws3.row_dimensions[row_idx].height = 22

# Sheet 2: Outreach templates
ws2 = wb.create_sheet("Plantillas Outreach")
ws2.column_dimensions["A"].width = 85

ws2["A1"] = "PLANTILLA DM INSTAGRAM"
ws2["A1"].font = Font(bold=True, size=14, color="FFD700")
ws2["A1"].fill = PatternFill("solid", fgColor="1a1a2e")
ws2.row_dimensions[1].height = 30

dm = (
    "Hola [NOMBRE] 👋\n\n"
    "Te escribo desde VISAI, una app de analisis de morfologia facial que recomienda "
    "cortes de pelo personalizados con IA.\n\n"
    "Lo que hacemos es exactamente lo que tu ya explicas a tus clientes: "
    "que corte le va a cada cara — pero digitalizado.\n\n"
    "La propuesta:\n"
    "- Tienes un codigo exclusivo (ej: [NOMBRE]VISAI)\n"
    "- Tus seguidores lo usan al hacer el analisis → ahorran 2€\n"
    "- Tu ganas 2€ por cada uso\n\n"
    "Sin coste, sin exclusividad, sin compromiso.\n\n"
    "Te envio mas info? 🙏"
)
ws2["A3"] = dm
ws2["A3"].alignment = Alignment(wrap_text=True, vertical="top")
ws2["A3"].font = Font(name="Calibri", size=11, color="FFFFFF")
ws2["A3"].fill = PatternFill("solid", fgColor="111111")
ws2.row_dimensions[3].height = 180

ws2["A6"] = "PLANTILLA EMAIL"
ws2["A6"].font = Font(bold=True, size=14, color="FFD700")
ws2["A6"].fill = PatternFill("solid", fgColor="1a1a2e")
ws2.row_dimensions[6].height = 30

email = (
    "Asunto: Colaboracion app VISAI — codigo de barbero para tus clientes\n\n"
    "Hola [NOMBRE],\n\n"
    "Me llamo Lucas, cofundador de VISAI (visaiapp.com). "
    "Hacemos analisis de morfologia facial con IA que recomienda el corte de pelo mas adecuado para cada cliente.\n\n"
    "He visto tu trabajo y creo que hay alineacion real: lo que nosotros hacemos digitalmente "
    "es lo que tu haces en consulta — analisis de cara y recomendacion personalizada.\n\n"
    "Propuesta concreta:\n"
    "- Codigo exclusivo tuyo (ej: [NOMBRE]VISAI)\n"
    "- Tus clientes lo usan → ahorran 2€ en el analisis\n"
    "- Tu ganas 2€ por cada uso, acumulable mensualmente\n\n"
    "Sin coste. Sin exclusividad. Sin compromiso minimo.\n\n"
    "Hacemos una llamada de 15 min para contarte?\n\n"
    "Saludos,\n"
    "Lucas — VISAI\n"
    "visaiapp.com"
)
ws2["A8"] = email
ws2["A8"].alignment = Alignment(wrap_text=True, vertical="top")
ws2["A8"].font = Font(name="Calibri", size=11, color="FFFFFF")
ws2["A8"].fill = PatternFill("solid", fgColor="111111")
ws2.row_dimensions[8].height = 300

wb.save("leads/barberos_influencers.xlsx")
print("Guardado: leads/barberos_influencers.xlsx")
print(f"{len(data)} influencers — {sum(1 for d in data if d[0]=='MEGA')} MEGA, {sum(1 for d in data if d[0]=='MID')} MID, {sum(1 for d in data if d[0]=='LOCAL')} LOCAL")

# Summary
with_email = [d[2] for d in data if d[6] and d[6] != "-"]
with_phone = [d[2] for d in data if d[7] and d[7] != "-"]
print(f"\nCon email ({len(with_email)}): {', '.join(with_email)}")
print(f"Con telefono ({len(with_phone)}): {', '.join(with_phone)}")
print(f"Solo DM: {len(data) - len(set(with_email + with_phone))}")
