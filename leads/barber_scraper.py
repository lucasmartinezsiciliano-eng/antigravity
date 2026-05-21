"""
barber_scraper.py — Generador de leads de barberías en España
=============================================================
Uso:
    python barber_scraper.py --steps 1,2,3,4 --output barberias_espana.xlsx --limit 1000
    python barber_scraper.py --steps 1 --output barberias_espana.xlsx
    python barber_scraper.py --steps 2,3,4 --input checkpoint.csv --output barberias_espana.xlsx

Pasos:
    1 — OpenStreetMap (Overpass API) — legal, sin key
    2 — Google Maps Places API o fallback DuckDuckGo
    3 — Instagram handles via búsqueda web
    4 — Emails desde webs de los negocios

Requiere:
    pip install -r requirements.txt

Variables de entorno opcionales:
    GOOGLE_MAPS_API_KEY — activa Google Maps Places API en paso 2
"""

import argparse
import json
import logging
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional
from urllib.parse import quote_plus, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from tqdm import tqdm

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
# StreamHandler con UTF-8 explícito para Windows (evita UnicodeEncodeError cp1252)
_stream_handler = logging.StreamHandler(sys.stdout)
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
_stream_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        _stream_handler,
        logging.FileHandler("barber_scraper.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------
TODAY = datetime.now().strftime("%Y-%m-%d")

CIUDADES_GRANDES = [
    "Madrid", "Barcelona", "Valencia", "Sevilla", "Zaragoza", "Málaga",
    "Murcia", "Palma", "Las Palmas", "Bilbao", "Alicante", "Córdoba",
    "Valladolid", "Vigo", "Gijón", "Granada", "Vitoria", "La Coruña",
    "Elche", "Santa Cruz de Tenerife", "Badalona", "Cartagena", "Terrassa",
    "Jerez de la Frontera", "Sabadell", "Móstoles", "Alcalá de Henares",
    "Fuenlabrada", "Almería", "Hospitalet", "San Sebastián", "Getafe",
    "Logroño", "Badajoz", "Huelva", "Salamanca", "Burgos", "Albacete",
    "Castellón", "Alcorcón", "Leganés", "Santander", "Pamplona", "Toledo",
    "Alcobendas", "Guadalajara", "Cádiz", "Marbella", "León",
]

HEADERS_BROWSER = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

EMAIL_REGEX = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
)

DOMINIOS_GENERICOS = {
    "gmail.com", "hotmail.com", "yahoo.com", "yahoo.es", "outlook.com",
    "live.com", "icloud.com", "me.com", "msn.com", "protonmail.com",
}

CHECKPOINT_INTERVAL = 100

INSTAGRAM_RE = re.compile(
    r"(?:https?://)?(?:www\.)?instagram\.com/([A-Za-z0-9_.]{1,30})/?",
    re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def _safe_get(url: str, timeout: int = 12, retries: int = 2, **kwargs) -> Optional[requests.Response]:
    """GET con reintentos y manejo silencioso de errores."""
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=HEADERS_BROWSER, timeout=timeout, **kwargs)
            r.raise_for_status()
            return r
        except Exception as exc:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                log.debug("GET fallido %s: %s", url, exc)
    return None


def _normalize_phone(raw: str) -> str:
    """Limpia un teléfono: quita espacios, puntos, guiones."""
    if not raw:
        return ""
    cleaned = re.sub(r"[\s.\-()]", "", raw)
    # Normaliza prefijo España
    if cleaned.startswith("0034"):
        cleaned = "+" + cleaned[2:]
    elif cleaned.startswith("34") and len(cleaned) >= 11:
        cleaned = "+" + cleaned
    elif cleaned.startswith("6") or cleaned.startswith("7") or cleaned.startswith("9"):
        cleaned = "+34" + cleaned
    return cleaned


def _extract_province_from_city(city: str) -> str:
    """Mapeo básico ciudad → provincia (ampliable)."""
    mapping = {
        "Madrid": "Madrid", "Barcelona": "Barcelona", "Valencia": "Valencia",
        "Sevilla": "Sevilla", "Zaragoza": "Zaragoza", "Málaga": "Málaga",
        "Murcia": "Murcia", "Palma": "Islas Baleares", "Las Palmas": "Las Palmas",
        "Bilbao": "Vizcaya", "Alicante": "Alicante", "Córdoba": "Córdoba",
        "Valladolid": "Valladolid", "Vigo": "Pontevedra", "Gijón": "Asturias",
        "Granada": "Granada", "Vitoria": "Álava", "La Coruña": "A Coruña",
        "Elche": "Alicante", "Santa Cruz de Tenerife": "Santa Cruz de Tenerife",
        "Badalona": "Barcelona", "Cartagena": "Murcia", "Terrassa": "Barcelona",
        "Jerez de la Frontera": "Cádiz", "Sabadell": "Barcelona",
        "Móstoles": "Madrid", "Alcalá de Henares": "Madrid",
        "Fuenlabrada": "Madrid", "Almería": "Almería",
        "Hospitalet": "Barcelona", "San Sebastián": "Guipúzcoa",
        "Getafe": "Madrid", "Logroño": "La Rioja", "Badajoz": "Badajoz",
        "Huelva": "Huelva", "Salamanca": "Salamanca", "Burgos": "Burgos",
        "Albacete": "Albacete", "Castellón": "Castellón",
        "Alcorcón": "Madrid", "Leganés": "Madrid", "Santander": "Cantabria",
        "Pamplona": "Navarra", "Toledo": "Toledo", "Alcobendas": "Madrid",
        "Guadalajara": "Guadalajara", "Cádiz": "Cádiz", "Marbella": "Málaga",
        "León": "León",
    }
    return mapping.get(city, "")


# ---------------------------------------------------------------------------
# Geocodificación inversa — Nominatim (OSM, gratuito, 1 req/s)
# ---------------------------------------------------------------------------

NOMINATIM_URL = "https://nominatim.openstreetmap.org/reverse"
_geo_cache: dict[tuple, dict] = {}  # (lat_r, lon_r) → {ciudad, provincia}

def _reverse_geocode(lat: float, lon: float) -> dict:
    """
    Devuelve {'ciudad': ..., 'provincia': ...} usando Nominatim.
    Agrupa por cuadrícula ~1km (2 decimales) para minimizar peticiones.
    Respeta el rate limit de 1 req/s con sleep().
    """
    lat_r = round(float(lat), 2)
    lon_r = round(float(lon), 2)
    key = (lat_r, lon_r)
    if key in _geo_cache:
        return _geo_cache[key]

    try:
        r = requests.get(
            NOMINATIM_URL,
            params={"lat": lat, "lon": lon, "format": "jsonv2", "addressdetails": 1},
            headers={"User-Agent": "barber-lead-scraper/1.0 (contact: barber@leads.es)"},
            timeout=10,
        )
        r.raise_for_status()
        addr = r.json().get("address", {})
        ciudad = (
            addr.get("city")
            or addr.get("town")
            or addr.get("village")
            or addr.get("municipality")
            or ""
        )
        # Nominatim devuelve la provincia como "state" o "county" en España
        provincia = addr.get("state") or addr.get("county") or ""
        # Quitar "Provincia de" / "Comunidad de" del inicio
        provincia = re.sub(r"^(?:Provincia de |Comunidad (?:de |Autónoma de )?|Región de )", "", provincia)
        result = {"ciudad": ciudad, "provincia": provincia}
    except Exception as exc:
        log.debug("Nominatim falló lat=%s lon=%s: %s", lat, lon, exc)
        result = {"ciudad": "", "provincia": ""}

    _geo_cache[key] = result
    time.sleep(1.1)  # Nominatim ToS: max 1 req/s
    return result


def _is_blank(val) -> bool:
    """True si el valor es None, NaN, o string vacío. Necesario porque pandas carga vacíos como float NaN."""
    import math
    if val is None:
        return True
    if isinstance(val, float) and math.isnan(val):
        return True
    return str(val).strip() == ""


def step_geocode(records: list) -> list:
    """
    Geocodificación en batch: agrupa coordenadas únicas (cuadrícula 10km),
    geocodifica cada grupo UNA VEZ, aplica a todos los registros del grupo.
    ~500-800 llamadas API para España entera → ~15-20 min vs 5h por registro.
    """
    log.info("=== PASO GEOCODIFICACIÓN: Nominatim batch reverse geocoding ===")

    pending = [
        r for r in records
        if (_is_blank(r.get("ciudad")) or _is_blank(r.get("provincia")))
        and not _is_blank(r.get("lat")) and not _is_blank(r.get("lon"))
    ]
    log.info("%d registros sin ciudad/provincia con coordenadas", len(pending))

    # Agrupar por cuadrícula 10km (1 decimal) para minimizar llamadas API
    from collections import defaultdict
    groups: dict[tuple, list] = defaultdict(list)
    for r in pending:
        try:
            key = (round(float(r["lat"]), 1), round(float(r["lon"]), 1))
            groups[key].append(r)
        except Exception:
            pass

    log.info("Cuadrículas únicas (10km): %d → estimado ~%d min", len(groups), len(groups) * 2 // 60)

    for (lat_r, lon_r), group_records in tqdm(groups.items(), desc="Geocodificando cuadrículas"):
        # Toma el centroide del grupo como representante
        lats = [float(r["lat"]) for r in group_records]
        lons = [float(r["lon"]) for r in group_records]
        lat_c = sum(lats) / len(lats)
        lon_c = sum(lons) / len(lons)
        try:
            geo = _reverse_geocode(lat_c, lon_c)
            for r in group_records:
                if _is_blank(r.get("ciudad")) and geo["ciudad"]:
                    r["ciudad"] = geo["ciudad"]
                if _is_blank(r.get("provincia")) and geo["provincia"]:
                    r["provincia"] = geo["provincia"]
                if not _is_blank(r.get("ciudad")) and _is_blank(r.get("provincia")):
                    r["provincia"] = _extract_province_from_city(r["ciudad"])
        except Exception:
            pass

    log.info("Geocodificación completada. Cuadrículas procesadas: %d", len(_geo_cache))
    return records


def _save_checkpoint(records: list, path: str):
    """Guarda checkpoint CSV."""
    if records:
        pd.DataFrame(records).to_csv(path, index=False, encoding="utf-8-sig")
        log.info("Checkpoint guardado: %d registros -> %s", len(records), path)


def _load_checkpoint(path: str) -> list:
    """Carga checkpoint CSV si existe."""
    if os.path.exists(path):
        df = pd.read_csv(path, encoding="utf-8-sig")
        log.info("Checkpoint cargado: %d registros desde %s", len(df), path)
        return df.to_dict("records")
    return []


def _deduplicate(records: list) -> list:
    """Elimina duplicados por nombre+ciudad."""
    seen = set()
    out = []
    for r in records:
        key = (str(r.get("nombre", "")).lower().strip(),
               str(r.get("ciudad", "")).lower().strip())
        if key not in seen:
            seen.add(key)
            out.append(r)
    return out


# ---------------------------------------------------------------------------
# PASO 1 — OpenStreetMap / Overpass API
# ---------------------------------------------------------------------------

OVERPASS_URL = "https://overpass-api.de/api/interpreter"

OVERPASS_QUERY = """
[out:json][timeout:300];
area["ISO3166-1"="ES"]->.spain;
(
  node["shop"="hairdresser"](area.spain);
  node["craft"="barber"](area.spain);
  node["shop"="barber"](area.spain);
  way["shop"="hairdresser"](area.spain);
  way["craft"="barber"](area.spain);
  way["shop"="barber"](area.spain);
  relation["shop"="hairdresser"](area.spain);
  relation["craft"="barber"](area.spain);
);
out center tags;
"""


def _parse_osm_element(el: dict) -> dict:
    tags = el.get("tags", {})
    # Coordenadas: node tiene lat/lon, way/relation tienen center
    lat = el.get("lat") or (el.get("center") or {}).get("lat") or ""
    lon = el.get("lon") or (el.get("center") or {}).get("lon") or ""

    name = tags.get("name") or tags.get("brand") or ""
    if not name:
        return {}

    phone = (
        tags.get("contact:phone")
        or tags.get("phone")
        or tags.get("contact:mobile")
        or tags.get("mobile")
        or ""
    )
    email = tags.get("contact:email") or tags.get("email") or ""
    website = tags.get("contact:website") or tags.get("website") or ""
    city = (
        tags.get("addr:city")
        or tags.get("addr:municipality")
        or tags.get("is_in:city")
        or ""
    )
    instagram = tags.get("contact:instagram") or tags.get("instagram") or ""
    # A veces el campo instagram es una URL completa
    if instagram and "instagram.com" in instagram:
        m = INSTAGRAM_RE.search(instagram)
        instagram = "@" + m.group(1) if m else instagram

    return {
        "nombre": name.strip(),
        "ciudad": city.strip(),
        "provincia": _extract_province_from_city(city.strip()),
        "telefono": _normalize_phone(phone),
        "email": email.lower().strip(),
        "instagram": instagram.strip(),
        "web": website.strip(),
        "fuente": "osm",
        "lat": lat,
        "lon": lon,
        "fecha": TODAY,
    }


def step1_osm(limit: int = 0) -> list:
    """Descarga barberías de España via Overpass API."""
    log.info("=== PASO 1: OpenStreetMap / Overpass API ===")
    try:
        log.info("Enviando query Overpass (puede tardar 1-3 min)...")
        r = requests.post(
            OVERPASS_URL,
            data={"data": OVERPASS_QUERY},
            timeout=360,
            headers={"User-Agent": "barber-lead-scraper/1.0 (contact: barber@leads.es)"},
        )
        r.raise_for_status()
        data = r.json()
    except Exception as exc:
        log.error("Overpass API falló: %s", exc)
        return []

    elements = data.get("elements", [])
    log.info("Overpass devolvió %d elementos", len(elements))

    records = []
    for el in tqdm(elements, desc="Parseando OSM"):
        rec = _parse_osm_element(el)
        if rec:
            records.append(rec)

    records = _deduplicate(records)
    log.info("Paso 1 completo: %d barberías únicas de OSM", len(records))

    if limit and len(records) > limit:
        records = records[:limit]

    return records


# ---------------------------------------------------------------------------
# PASO 2 — Google Maps Places API + fallback DuckDuckGo/Bing
# ---------------------------------------------------------------------------

def _step2_googlemaps_api(ciudad: str, api_key: str) -> list:
    """Busca barberías en una ciudad usando Google Maps Places API."""
    try:
        import googlemaps
        gmaps = googlemaps.Client(key=api_key)
    except ImportError:
        log.warning("Librería googlemaps no instalada: pip install googlemaps")
        return []

    records = []
    queries = [f"barbería {ciudad}", f"barber shop {ciudad}", f"peluquería hombre {ciudad}"]

    for query in queries:
        try:
            results = gmaps.places(query=query, language="es", region="es")
            for place in results.get("results", []):
                name = place.get("name", "")
                if not name:
                    continue

                # Detalles adicionales
                place_id = place.get("place_id", "")
                details = {}
                if place_id:
                    try:
                        det = gmaps.place(
                            place_id=place_id,
                            fields=["name", "formatted_phone_number",
                                    "website", "formatted_address"],
                            language="es",
                        )
                        details = det.get("result", {})
                    except Exception:
                        pass

                address = details.get("formatted_address") or place.get("formatted_address", "")
                # Extrae ciudad de la dirección si no la tenemos
                phone = details.get("formatted_phone_number", "")
                website = details.get("website", "")
                loc = place.get("geometry", {}).get("location", {})

                records.append({
                    "nombre": name.strip(),
                    "ciudad": ciudad,
                    "provincia": _extract_province_from_city(ciudad),
                    "telefono": _normalize_phone(phone),
                    "email": "",
                    "instagram": "",
                    "web": website.strip(),
                    "fuente": "google_maps",
                    "lat": loc.get("lat", ""),
                    "lon": loc.get("lng", ""),
                    "fecha": TODAY,
                })
            time.sleep(0.5)
        except Exception as exc:
            log.debug("Google Maps query '%s' falló: %s", query, exc)

    return records


def _step2_duckduckgo_fallback(ciudad: str) -> list:
    """Fallback: busca barberías via DuckDuckGo Instant Answer API."""
    records = []
    # DuckDuckGo no tiene Places API, usamos búsqueda de texto como señal
    # Se parsea lo que se puede de los snippets
    queries = [
        f"barbería {ciudad} teléfono web",
        f"barber shop {ciudad} contacto",
    ]
    for query in queries:
        try:
            url = f"https://api.duckduckgo.com/?q={quote_plus(query)}&format=json&no_html=1&skip_disambig=1"
            r = _safe_get(url, timeout=10)
            if not r:
                continue
            data = r.json()
            # RelatedTopics puede tener resultados
            for topic in data.get("RelatedTopics", [])[:5]:
                text = topic.get("Text", "")
                if text and len(text) > 10:
                    # Intenta extraer un nombre de negocio del snippet
                    name_match = re.match(r"^([^–\-|]+)", text)
                    name = name_match.group(1).strip()[:80] if name_match else ""
                    if name and len(name) > 3:
                        records.append({
                            "nombre": name,
                            "ciudad": ciudad,
                            "provincia": _extract_province_from_city(ciudad),
                            "telefono": "",
                            "email": "",
                            "instagram": "",
                            "web": topic.get("FirstURL", ""),
                            "fuente": "duckduckgo",
                            "lat": "",
                            "lon": "",
                            "fecha": TODAY,
                        })
            time.sleep(1.5)
        except Exception as exc:
            log.debug("DuckDuckGo query '%s' falló: %s", query, exc)

    return records


def _step2_bing_maps(ciudad: str) -> list:
    """Busca barberías en Bing Maps (sin key, HTML scraping limitado)."""
    records = []
    queries = [f"barbería {ciudad}", f"barber {ciudad}"]
    for query in queries:
        try:
            url = f"https://www.bing.com/search?q={quote_plus(query + ' horario teléfono')}&setlang=es"
            r = _safe_get(url, timeout=12)
            if not r:
                continue
            soup = BeautifulSoup(r.text, "lxml")

            # Bing muestra Local Pack con li.b_algo o div.b_lcl
            for card in soup.select("li.b_algo")[:8]:
                title_el = card.select_one("h2 a")
                if not title_el:
                    continue
                name = title_el.get_text(strip=True)
                href = title_el.get("href", "")

                # Teléfono en snippet
                snippet = card.get_text(" ", strip=True)
                phone_match = re.search(r"(\+34\s?)?[679]\d{2}[\s.\-]?\d{3}[\s.\-]?\d{3}", snippet)
                phone = phone_match.group(0) if phone_match else ""

                if len(name) > 3:
                    records.append({
                        "nombre": name[:120],
                        "ciudad": ciudad,
                        "provincia": _extract_province_from_city(ciudad),
                        "telefono": _normalize_phone(phone),
                        "email": "",
                        "instagram": "",
                        "web": href if href.startswith("http") else "",
                        "fuente": "bing_search",
                        "lat": "",
                        "lon": "",
                        "fecha": TODAY,
                    })
            time.sleep(2)
        except Exception as exc:
            log.debug("Bing query '%s' falló: %s", query, exc)
    return records


def step2_maps(existing: list, limit: int = 0) -> list:
    """Enriquece con Google Maps o fallback para ciudades grandes."""
    log.info("=== PASO 2: Google Maps / Fallback ===")
    api_key = os.environ.get("GOOGLE_MAPS_API_KEY", "")
    if api_key:
        log.info("Google Maps API key detectada — usando Places API")
    else:
        log.warning("GOOGLE_MAPS_API_KEY no definida — usando Bing/DuckDuckGo como fallback")

    all_records = list(existing)
    # Ciudades ya cubiertas en paso 1
    ciudades_ya = {r.get("ciudad", "").lower() for r in existing if r.get("ciudad")}

    new_records = []
    for ciudad in tqdm(CIUDADES_GRANDES, desc="Ciudades"):
        try:
            if api_key:
                recs = _step2_googlemaps_api(ciudad, api_key)
            else:
                # Combinamos ambos fallbacks
                recs = _step2_bing_maps(ciudad)
                recs += _step2_duckduckgo_fallback(ciudad)

            new_records.extend(recs)
            log.debug("Ciudad %s: +%d registros", ciudad, len(recs))
            time.sleep(1)
        except Exception as exc:
            log.warning("Ciudad %s falló: %s", ciudad, exc)

    combined = all_records + new_records
    combined = _deduplicate(combined)
    log.info("Paso 2 completo: %d registros totales", len(combined))

    if limit and len(combined) > limit:
        combined = combined[:limit]

    return combined


# ---------------------------------------------------------------------------
# PASO 3 — Instagram handles via búsqueda web
# ---------------------------------------------------------------------------

MAX_INSTAGRAM_QUERIES_PER_HOUR = 500
_instagram_query_count = 0
_instagram_hour_start = time.time()


def _rate_limit_instagram():
    global _instagram_query_count, _instagram_hour_start
    _instagram_query_count += 1
    elapsed = time.time() - _instagram_hour_start
    if elapsed > 3600:
        _instagram_query_count = 1
        _instagram_hour_start = time.time()
    if _instagram_query_count >= MAX_INSTAGRAM_QUERIES_PER_HOUR:
        wait = 3600 - elapsed
        log.info("Rate limit Instagram: esperando %.0f segundos...", wait)
        time.sleep(max(wait, 0))
        _instagram_query_count = 0
        _instagram_hour_start = time.time()


def _find_instagram_bing(nombre: str, ciudad: str) -> str:
    """Busca el handle de Instagram via Bing."""
    query = f'"{nombre}" "{ciudad}" site:instagram.com'
    try:
        url = f"https://www.bing.com/search?q={quote_plus(query)}&setlang=es"
        r = _safe_get(url, timeout=10)
        if not r:
            return ""
        soup = BeautifulSoup(r.text, "lxml")
        for a in soup.select("a[href*='instagram.com']"):
            href = a.get("href", "")
            m = INSTAGRAM_RE.search(href)
            if m:
                handle = m.group(1)
                # Filtra handles genéricos de Instagram
                if handle.lower() not in {"p", "explore", "accounts", "reel", "stories", "tv", ""}:
                    return "@" + handle
        # Busca en texto de la página también
        m = INSTAGRAM_RE.search(r.text)
        if m:
            handle = m.group(1)
            if handle.lower() not in {"p", "explore", "accounts", "reel", "stories", "tv", ""}:
                return "@" + handle
    except Exception as exc:
        log.debug("Instagram Bing '%s' falló: %s", nombre, exc)
    return ""


def _find_instagram_duckduckgo(nombre: str, ciudad: str) -> str:
    """Busca el handle de Instagram via DuckDuckGo."""
    query = f'"{nombre}" "{ciudad}" instagram'
    try:
        url = f"https://api.duckduckgo.com/?q={quote_plus(query)}&format=json&no_html=1"
        r = _safe_get(url, timeout=10)
        if not r:
            return ""
        text = r.text
        m = INSTAGRAM_RE.search(text)
        if m:
            handle = m.group(1)
            if handle.lower() not in {"p", "explore", "accounts", "reel", "stories", "tv", ""}:
                return "@" + handle
    except Exception as exc:
        log.debug("Instagram DDG '%s' falló: %s", nombre, exc)
    return ""


def _find_instagram_from_web(website: str) -> str:
    """Busca enlace a Instagram en la web del negocio."""
    if not website:
        return ""
    try:
        r = _safe_get(website, timeout=10)
        if not r:
            return ""
        m = INSTAGRAM_RE.search(r.text)
        if m:
            handle = m.group(1)
            if handle.lower() not in {"p", "explore", "accounts", "reel", "stories", "tv", ""}:
                return "@" + handle
    except Exception as exc:
        log.debug("Instagram desde web '%s' falló: %s", website, exc)
    return ""


def step3_instagram(records: list) -> list:
    """Busca handles de Instagram para registros que no lo tienen."""
    log.info("=== PASO 3: Instagram handles ===")
    sin_instagram = [r for r in records if not r.get("instagram")]
    log.info("%d registros sin Instagram de %d totales", len(sin_instagram), len(records))

    updated = {id(r): r for r in records}

    for rec in tqdm(sin_instagram, desc="Instagram"):
        nombre = rec.get("nombre", "")
        ciudad = rec.get("ciudad", "")
        website = rec.get("web", "")

        if not nombre:
            continue

        instagram = ""

        # 1) Busca en la web propia si hay
        if website:
            instagram = _find_instagram_from_web(website)

        # 2) Bing search
        if not instagram and ciudad:
            _rate_limit_instagram()
            instagram = _find_instagram_bing(nombre, ciudad)
            time.sleep(2)

        # 3) DuckDuckGo
        if not instagram and ciudad:
            _rate_limit_instagram()
            instagram = _find_instagram_duckduckgo(nombre, ciudad)
            time.sleep(2)

        if instagram:
            updated[id(rec)]["instagram"] = instagram

    result = list(updated.values())
    con_instagram = sum(1 for r in result if r.get("instagram"))
    log.info("Paso 3 completo: %d registros con Instagram", con_instagram)
    return result


# ---------------------------------------------------------------------------
# PASO 4 — Emails desde webs
# ---------------------------------------------------------------------------

PATHS_CONTACTO = ["", "/contacto", "/contact", "/sobre-nosotros", "/about",
                  "/quienes-somos", "/info", "/contactar"]


def _extract_emails_from_html(html: str, domain: str) -> list:
    """Extrae emails válidos del HTML, filtrando dominios genéricos."""
    found = EMAIL_REGEX.findall(html)
    result = []
    for email in found:
        email = email.lower().strip(".,;:")
        parts = email.split("@")
        if len(parts) != 2:
            continue
        email_domain = parts[1]
        # Prefiere emails del mismo dominio del negocio
        if email_domain in DOMINIOS_GENERICOS:
            result.append((email, 0))  # prioridad baja
        elif domain and domain in email_domain:
            result.append((email, 2))  # prioridad alta: mismo dominio
        else:
            result.append((email, 1))  # prioridad media: otro dominio propio
    # Ordena por prioridad descendente
    result.sort(key=lambda x: -x[1])
    # Devuelve únicos respetando orden
    seen = set()
    out = []
    for email, _ in result:
        if email not in seen:
            seen.add(email)
            out.append(email)
    return out


def _get_domain(url: str) -> str:
    try:
        return urlparse(url).netloc.replace("www.", "")
    except Exception:
        return ""


def _fetch_emails_from_website(website: str) -> str:
    """Descarga varias páginas del sitio y extrae el mejor email."""
    if not website:
        return ""

    # Normaliza URL base
    if not website.startswith("http"):
        website = "https://" + website
    base = website.rstrip("/")
    domain = _get_domain(base)

    all_emails = []
    for path in PATHS_CONTACTO[:6]:  # Limita a 6 rutas para no ser lento
        url = base + path
        try:
            r = _safe_get(url, timeout=10, retries=1)
            if r:
                emails = _extract_emails_from_html(r.text, domain)
                all_emails.extend(emails)
                if emails:
                    break  # Encontró emails, no sigue
        except Exception:
            pass
        time.sleep(0.3)

    # Devuelve el primer email de mayor prioridad
    return all_emails[0] if all_emails else ""


def step4_emails(records: list) -> list:
    """Extrae emails de las webs de los negocios."""
    log.info("=== PASO 4: Emails desde webs ===")
    con_web = [r for r in records if r.get("web") and not r.get("email")]
    log.info("%d registros con web sin email de %d totales", len(con_web), len(records))

    updated = {id(r): r for r in records}

    for rec in tqdm(con_web, desc="Emails desde web"):
        website = rec.get("web", "")
        if not website:
            continue
        try:
            email = _fetch_emails_from_website(website)
            if email:
                updated[id(rec)]["email"] = email
        except Exception as exc:
            log.debug("Email desde web '%s' falló: %s", website, exc)

    result = list(updated.values())
    con_email = sum(1 for r in result if r.get("email"))
    log.info("Paso 4 completo: %d registros con email", con_email)
    return result


# ---------------------------------------------------------------------------
# PASO 5 — Output Excel
# ---------------------------------------------------------------------------

COLUMNAS = [
    "nombre", "ciudad", "provincia", "telefono", "email",
    "instagram", "web", "fuente", "lat", "lon", "fecha",
]


def step5_excel(records: list, output_path: str):
    """Genera el Excel final con formato."""
    log.info("=== PASO 5: Exportando Excel ===")
    df = pd.DataFrame(records, columns=COLUMNAS)

    # Limpieza final
    df = df.fillna("")
    df["nombre"] = df["nombre"].str.strip()
    df["ciudad"] = df["ciudad"].str.strip()
    df = df[df["nombre"] != ""]
    df = df.drop_duplicates(subset=["nombre", "ciudad"])
    df = df.sort_values(["provincia", "ciudad", "nombre"])
    df = df.reset_index(drop=True)

    log.info("Exportando %d registros a %s", len(df), output_path)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Barberías")

        ws = writer.sheets["Barberías"]

        # Estilo cabecera
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Anchos de columna automáticos
        col_widths = {
            "nombre": 35, "ciudad": 18, "provincia": 18,
            "telefono": 16, "email": 32, "instagram": 22,
            "web": 40, "fuente": 14, "lat": 12, "lon": 12, "fecha": 12,
        }
        for i, col in enumerate(COLUMNAS, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = col_widths.get(col, 15)

        # Freeze primera fila
        ws.freeze_panes = "A2"

        # Auto-filtro
        ws.auto_filter.ref = ws.dimensions

    log.info("Excel generado: %s", output_path)
    return df


# ---------------------------------------------------------------------------
# CLI principal
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description="Genera leads de barberías en España",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--steps",
        default="1,2,3,4",
        help="Pasos a ejecutar separados por coma (default: 1,2,3,4)",
    )
    parser.add_argument(
        "--output",
        default="barberias_espana.xlsx",
        help="Ruta del Excel de salida (default: barberias_espana.xlsx)",
    )
    parser.add_argument(
        "--input",
        default=None,
        help="CSV de checkpoint para continuar desde un estado previo",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Máximo de registros a procesar (0 = sin límite)",
    )
    parser.add_argument(
        "--checkpoint",
        default="checkpoint.csv",
        help="Ruta del CSV checkpoint (default: checkpoint.csv)",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Carga checkpoint automáticamente sin preguntar",
    )
    parser.add_argument(
        "--fresh",
        action="store_true",
        help="Ignora checkpoint existente y empieza desde cero",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    steps = [int(s.strip()) for s in args.steps.split(",") if s.strip().isdigit()]
    log.info("Iniciando barber_scraper | pasos=%s | limit=%d | output=%s",
             steps, args.limit, args.output)

    # Carga estado previo si se proporcionó --input
    records = []
    if args.input:
        records = _load_checkpoint(args.input)
    elif os.path.exists(args.checkpoint) and not args.fresh:
        if args.resume or not sys.stdin.isatty():
            records = _load_checkpoint(args.checkpoint)
        else:
            log.info("Checkpoint previo encontrado en %s", args.checkpoint)
            resp = input("¿Cargar checkpoint? [s/N]: ").strip().lower()
            if resp == "s":
                records = _load_checkpoint(args.checkpoint)

    # PASO 1 — OSM
    if 1 in steps:
        records_osm = step1_osm(limit=args.limit)
        # Merge: los del checkpoint tienen prioridad
        existing_keys = {(r.get("nombre", "").lower(), r.get("ciudad", "").lower())
                         for r in records}
        for r in records_osm:
            key = (r.get("nombre", "").lower(), r.get("ciudad", "").lower())
            if key not in existing_keys:
                records.append(r)
        records = _deduplicate(records)
        _save_checkpoint(records, args.checkpoint)

    # PASO 2 — Maps
    if 2 in steps:
        records = step2_maps(records, limit=args.limit)
        _save_checkpoint(records, args.checkpoint)

    # PASO 3 — Instagram
    if 3 in steps:
        records = step3_instagram(records)
        _save_checkpoint(records, args.checkpoint)

    # PASO 4 — Emails
    if 4 in steps:
        records = step4_emails(records)
        _save_checkpoint(records, args.checkpoint)

    # PASO 6 — Geocodificación inversa (ciudad + provincia desde lat/lon)
    if 6 in steps:
        records = step_geocode(records)
        _save_checkpoint(records, args.checkpoint)

    # PASO 5 — Excel (siempre al final)
    if records:
        df = step5_excel(records, args.output)
        log.info("=" * 60)
        log.info("FINALIZADO: %d barberías exportadas a %s", len(df), args.output)
        log.info("Con teléfono:  %d", df["telefono"].astype(bool).sum())
        log.info("Con email:     %d", df["email"].astype(bool).sum())
        log.info("Con Instagram: %d", df["instagram"].astype(bool).sum())
        log.info("Con web:       %d", df["web"].astype(bool).sum())
        log.info("Con ciudad:    %d", df["ciudad"].astype(bool).sum())
        log.info("Con provincia: %d", df["provincia"].astype(bool).sum())
        log.info("=" * 60)
    else:
        log.warning("No se obtuvieron registros. Revisa la conexión o los pasos seleccionados.")


if __name__ == "__main__":
    main()
